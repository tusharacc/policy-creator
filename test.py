import CH_data as CH
import time,traceback,names
import user_detail as user
import selenium.common.exceptions
from splinter import Browser
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from bs4 import BeautifulSoup

def checkPageTransition(old,new,msg):
    if old == new:
        raise ValueError(msg)

def flipTheValue(orig,ind,question_level):
    
    if question_level == 'P':
        return orig
    else:
        if ind == 'N':
            return int(not bool(orig))
        else:
            return orig
            
test_condition = input("Type N for negative scenario and P for positive scenario :  ").upper()

#the list will contain all the values from excel
input_test_policies = []
wb = load_workbook('Test-Case.xlsm')
ws = wb['Sheet1']

for i in range(2,3):
    CH.value_read['state'] = ws['A'+str(i)].value
    CH.value_read['business_segment'] = ws['B'+str(i)].value
    CH.value_read['business_type'] = ws['C'+str(i)].value
    CH.value_read['curr_coverage'] = ws['D'+str(i)].value
    CH.value_read['business_ownership'] = ws['G'+str(i)].value
    CH.value_read['business_start_date'] = ws['H'+str(i)].value
    CH.value_read['employees_count'] = ws['J'+str(i)].value
    CH.value_read['annual_payroll'] = ws['K'+str(i)].value
    CH.value_read['annual_gross_sales'] = ws['L'+str(i)].value
    CH.value_read['footage'] = ws['M'+str(i)].value
    CH.value_read['address_line'] = ws['S'+str(i)].value
    CH.value_read['city'] = ws['U'+str(i)].value
    CH.value_read['zip_code'] = ws['W'+str(i)].value
    #print (CH.value_read)
    input_test_policies.append(dict(CH.value_read))
    #print (input_test_policies)

#print (input_test_policies)    
#Constants used in the program
COVERAGE_ID = ['product_codes__general_liability',
               'product_codes__professional_liability',
               'product_codes__workers_compensation',
               'product_codes__commercial_auto',
               'product_codes__bop']
COVERAGE_XPATH = {
    'gl':'//*[@id="field_for_product_codes__general_liability"]/label',
    'pl':'//*[@id="field_for_product_codes__professional_liability"]/label',
    'wc':'//*[@id="field_for_product_codes__workers_compensation"]/label',
    'ca':'//*[@id="field_for_product_codes__commercial_auto"]/label',
    'bp':'//*[@id="field_for_product_codes__bop"]/label'
}

EMAIL = 'bill.clinton@whitehouse.com'
ADDITIONAL_QUESTIONS = ['Personal Training (Health And Fitness)']

for policy in input_test_policies:
    COMPANY_NAME = names.get_full_name()
    print (COMPANY_NAME)
    FIRST_NAME = names.get_first_name()
    LAST_NAME = names.get_last_name()
    
    curr_url = 'https://'+user.USERNAME+':'+user.PASSWORD+'@psc-chubb-sit.coverhound.us/business-insurance'
    first_pass = True
    browser = Browser('phantomjs')
    driver = webdriver.PhantomJS()
    driver.delete_all_cookies()
    driver.start_session(DesiredCapabilities.PHANTOMJS)
    driver.implicitly_wait(10)
    driver.set_window_size(1120, 550)
    try:
        #HOME PAGE
        driver.get(curr_url)
        driver.execute_script('window.localStorage.clear();')
        driver.find_element_by_xpath('//*[@id="cm-quote-form"]/div[1]/div[1]/div[1]/div[1]/input').send_keys(CH.states[policy['state']])
        # select_state = Select(driver.find_element_by_id('state_abbrev'))
        # select_state.select_by_visible_text(CH.states[policy['state']])
        time.sleep(5)
        driver.find_element_by_xpath('//*[@id="cm-quote-form"]/div[1]/div[2]/div[1]/div[1]/input').send_keys(policy['business_type'])
        # select_business_segment = Select(driver.find_element_by_id('business_segment_id'))
        # select_business_segment.select_by_visible_text(policy['business_segment'])
        # time.sleep(5)
        # select_business_type = Select(driver.find_element_by_id('business_type_id'))
        # select_business_type.select_by_visible_text(policy['business_type'])
        
        driver.save_screenshot(COMPANY_NAME+ '_home_page_screenshot.png')
        driver.find_element_by_xpath('//*[@id="cm-quote-form"]/div[1]/button/span/span').click()
        time.sleep(10)
        #wait = WebDriverWait(driver, 10).until( EC.element_to_be_clickable((By.ID, 'product_codes__bop')))
        #print (driver.current_url)
        checkPageTransition(curr_url,driver.current_url,'Error in Home Page')
        
        for _type in COVERAGE_ID:
            driver.execute_script("document.getElementById('"+_type+"').checked = false")
        
        #driver.save_screenshot(COMPANY_NAME+ '_BusinessInfo_screenshot.png')
        
        if driver.current_url == 'https://'+user.USERNAME+':'+user.PASSWORD+'@psc-chubb-sit.coverhound.us/business-insurance/business-info':
            curr_url = driver.current_url
            driver.find_element_by_xpath(COVERAGE_XPATH[policy['curr_coverage']]).click()
            comp_name = driver.find_element_by_id('business_name')
            comp_name.clear()
            comp_name.send_keys(COMPANY_NAME)
            email = driver.find_element_by_id('email')
            email.clear()
            email.send_keys((Keys.CONTROL, "a"))
            email.send_keys(EMAIL)
            driver.save_screenshot(COMPANY_NAME+ '_business_info_screenshot.png')
            driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/form/div[1]/div/div/button').click()
            #print (driver.current_url)
            time.sleep(10)
            checkPageTransition(curr_url,driver.current_url,'Error in BusinessInfo')
            
        #BUSINESS OPERATION
        if driver.current_url == 'https://'+user.USERNAME+':'+user.PASSWORD+'@psc-chubb-sit.coverhound.us/business-insurance/business-operations':
            curr_url = driver.current_url
            select_business_structure = Select(driver.find_element_by_xpath('//*[@id="CH_024"]'))
            select_business_structure.select_by_visible_text(policy['business_ownership'])
            driver.find_element_by_id('CH_105').send_keys('04/04/2016')
            if first_pass:
                driver.find_element_by_xpath('//*[@id="field_for_CH_026"]/div[1]/div[2]/label').click()
            number_of_employees = driver.find_element_by_id('CH_029')
            number_of_employees.clear()
            number_of_employees.send_keys(policy['employees_count'])
            payroll = driver.find_element_by_id('CH_030')
            payroll.clear()
            payroll.send_keys(policy['annual_payroll'])
            projected_sales = driver.find_element_by_id('CH_031')
            projected_sales.clear()
            projected_sales.send_keys(policy['annual_gross_sales'])
            square_footage = driver.find_element_by_id('CH_032')
            square_footage.clear()
            square_footage.send_keys(policy['footage'])
            driver.save_screenshot(COMPANY_NAME+ '_business_operation_screenshot.png')
            driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/form/div[1]/div/div/button').click()
            #print (driver.current_url)
            time.sleep(10)
            checkPageTransition(curr_url,driver.current_url,'Error in BusinessOp')
            
        if driver.current_url == 'https://'+user.USERNAME+':'+user.PASSWORD+'@psc-chubb-sit.coverhound.us/business-insurance/contact':
            curr_url = driver.current_url
            select_title = Select(driver.find_element_by_id('title'))
            select_title.select_by_visible_text('Dr.')
            f_name = driver.find_element_by_id('first_name')
            f_name.clear()
            f_name.send_keys(FIRST_NAME)
            l_name = driver.find_element_by_id('last_name')
            l_name.clear()
            l_name.send_keys(LAST_NAME)
            ph_num = driver.find_element_by_id('telephone')
            ph_num.clear()
            ph_num.send_keys('8888888888')
            time.sleep(1)
            confirm_email = driver.find_element_by_id('email').get_attribute('value').encode('utf-8')
            address = driver.find_element_by_xpath('//*[@id="field_for_CH_020"]/div[1]/div[1]/input')
            address.clear()
            address.send_keys(policy['address_line'])
            time.sleep(1)
            city_text = driver.find_element_by_id('CH_037')
            city_text.clear()
            city_text.send_keys(policy['city'])
            time.sleep(1)
            select_insured_state = Select(driver.find_element_by_id('CH_018'))
            select_insured_state.select_by_visible_text(policy['state'])
            zip_text = driver.find_element_by_id('CH_038')
            zip_text.clear()
            zip_text.send_keys(policy['zip_code'])
            select_other_address = Select(driver.find_element_by_id('CH_027'))
            select_other_address.select_by_visible_text('No Additional Locations')
            driver.save_screenshot(COMPANY_NAME+ '_contact_info_screenshot.png')
            driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/form/div/div[1]/div/div/button').click()
            #print (driver.current_url)
            time.sleep(10)
            checkPageTransition(curr_url,driver.current_url,'Error in ContactInfo')
            
        if driver.current_url == 'https://'+user.USERNAME+':'+user.PASSWORD+'@psc-chubb-sit.coverhound.us/business-insurance/coverage-detail/bop':   
            curr_url = driver.current_url
            html =  driver.execute_script("return document.documentElement.outerHTML")
            soup = BeautifulSoup(html, 'html.parser')
            questions_div = soup.find_all("div", class_="question")
            i = 0
            for div in questions_div:
                i += 1
                label = div.find('label')
                question = label.text
                question_id = label['for']
                #print (question)
                try:
                    question_level,answer = CH.question_list[question]
                    #print (question_level + answer)
                    if answer == 'No':
                        driver.find_element_by_css_selector("label[for='"+question_id+"_"+str(flipTheValue(1,test_condition,question_level))+"']").click()
                        #print ("label[for='"+question_id+"_"+str(flipTheValue(1,test_condition,question_level))+"']")
                        #driver.execute_script("document.getElementById('"+question_id+"_1').checked = true")
                        #print ("document.getElementById('"+question_id+"_"+str(flipTheValue(1,test_condition,question_level))+"').checked = true")
                    elif answer == 'Yes':
                        driver.find_element_by_css_selector("label[for='"+question_id+"_"+str(flipTheValue(0,test_condition,question_level))+"']").click()
                        #print ("label[for='"+question_id+"_"+str(flipTheValue(0,test_condition,question_level))+"']")
                        #driver.execute_script("document.getElementById('"+question_id+"_0').checked = true")
                        #print ("document.getElementById('"+question_id+"_"+str(flipTheValue(0,test_condition,question_level))+"').checked = true")
                except KeyError:
                    if question == 'When would you like your coverage to start?':
                        pass
                    elif question == 'Does your business provide any of the following services? (Please select all that apply.)':
                        driver.find_element_by_xpath('//*[@id="field_for_CH_323__1122"]/label').click()
                    # else:
                    #     driver.find_element_by_css_selector("label[for='"+question_id+"_1']").click()
                
                #driver.save_screenshot(COMPANY_NAME+'_'+str(i)+'_'+question_id+'_coverage_detail_screenshot.png')  
            
            time.sleep(2)
            driver.save_screenshot(COMPANY_NAME+'_coverage_detail_screenshot.png')  
            driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/form/div/div[1]/div/div/button').click()
                                          
            time.sleep(60)
            checkPageTransition(curr_url,driver.current_url,'Error in Coverage')
            
            no_quote=  None
            quote = None
            try:
                no_quote = driver.find_element_by_class_name( "cm-no-quotes-yellow-light")
                quote = driver.find_element_by_class_name( "cm-product-quotes")
            except selenium.common.exceptions.NoSuchElementException:
                if no_quote == None:
                    driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/form/div/div[1]/div/div/div/div[2]/div[1]/div[2]/div[1]/a/span').click()
                    time.sleep(2)
                    driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/form/div/div[1]/div/div/button/span').click()
                    time.sleep(10)
                    driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/div/div/div[4]/div/a/span/span').click()
                    time.sleep(10)
                    #driver.find_element_by_class_name('cm-application-summary-acknowledgement-condition-checkbox').click()
                    driver.find_element_by_id('confirmation_name').send_keys(FIRST_NAME+' '+LAST_NAME)
                    driver.find_element_by_xpath(' //*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/div/div/div[5]/form/div[2]/div/button/span/span').click()
                    time.sleep(10)               
                    select_insured_state_checkout_page = Select(driver.find_element_by_id('address_state'))
                    select_insured_state_checkout_page.select_by_visible_text(policy['state'])
                    #driver.execute_script("document.getElementByName('cardnumber')[0].value = 5555 5555 5555 4444")
                    driver.switch_to.frame(driver.find_element_by_name('__privateStripeFrame4'))
                    driver.find_element_by_name('cardnumber').send_keys(5555)
                    #card_number = driver.find_elements_by_css_selector("#cardNumber div.__PrivateStripeElement input.__PrivateStripeElement-input")
                    # print (card_number)
                    # card_number[0].click()
                    # card_number[0].send_keys('5555')
                    #driver.find_element_by_xpath("//input[@name='cardnumber']").send_keys('5555555555554444')
                    #driver.find_element_by_xpath('//*[@id="root"]/form/span/label/input').send_keys('04/19')
                    #driver.find_element_by_xpath('//*[@id="root"]/form/span/label/input').send_keys('678')
                    #driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/div/div/section[2]/form/div[6]/button').click()
                    #time.sleep(30)
                    #policy_num = driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/div[2]/div/div/div[2]/ul/li/div/div/p[1]/span[2]').getText()
                    driver.save_screenshot(COMPANY_NAME+'_policy_num.png') 
                    #print (policy_num)
            #f = open('html_source.py','w')
            #f.write(html)
            #f.close()//*[@id="root"]/form/span/label/input
            # if first_pass:
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_327"]/div[1]/div[2]/label').click()
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_300"]/div[1]/div[2]/label').click()
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_301"]/div[1]/div[2]/label').click()
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_302"]/div[1]/div[2]/label').click()
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_303"]/div[1]/div[1]/label').click()
            #     driver.find_element_by_xpath('//*[@id="field_for_CH_304"]/div[1]/div[2]/label').click()
            #     if ADDITIONAL_QUESTIONS.count(policy['business_type']) > 0:
            #         driver.find_element_by_xpath('//*[@id="field_for_CH_322"]/div[1]/div[1]/label').click()
            #         driver.find_element_by_xpath('//*[@id="field_for_CH_323__1122"]/label').click()
            #     driver.save_screenshot(COMPANY_NAME+'_coverage_detail_screenshot.png')
            #     driver.find_element_by_xpath('//*[@id="commercial-app"]/div/div[2]/div[2]/div/form/div/div[1]/div/div/button').click()
            # print (driver.current_url)
            # time.sleep(60)
            # checkPageTransition(curr_url,driver.current_url,'Error in Coverage')
        
        driver.save_screenshot(COMPANY_NAME+'_success_screenshot.png')
        
    except Exception as e:
        print (e)
        print (driver.current_url)
        driver.save_screenshot(COMPANY_NAME +'_error_screenshot.png')
        traceback.print_exc()
    finally:
        driver.close()
        driver.quit()